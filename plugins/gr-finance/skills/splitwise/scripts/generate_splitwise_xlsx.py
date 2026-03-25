#!/usr/bin/env python3
"""Generate Splitwise-ready xlsx from structured expense data passed as JSON via stdin.

Usage:
    echo '<json>' | python3 generate_splitwise_xlsx.py
"""
import json
import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL    = PatternFill("solid", fgColor="2F5496")
TOTAL_FONT     = Font(bold=True, size=11)
TOTAL_FILL     = PatternFill("solid", fgColor="D6E4F0")
TITLE_FONT     = Font(bold=True, size=13, color="FFFFFF")
TITLE_FILL     = PatternFill("solid", fgColor="1F3864")
SUBTITLE_FONT  = Font(bold=True, size=11, color="2F5496")
ITEM_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
ITEM_HDR_FILL  = PatternFill("solid", fgColor="4472C4")
SUB_FILL       = PatternFill("solid", fgColor="D6E4F0")
SUB_FONT       = Font(bold=True, size=10)
MONEY_FMT      = '$#,##0.00'
BORDER         = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)


def hdr(ws, cols, row=1):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER


def style_data(ws, r1, r2, max_col, money_cols=None):
    for r in range(r1, r2 + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if money_cols and c in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal='right')


def auto_width(ws, max_col):
    for c in range(1, max_col + 1):
        w = max((len(str(cell.value)) for row in ws.iter_rows(min_col=c, max_col=c)
                 for cell in row if cell.value), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(w + 3, 32)


def total_row(ws, row, num_cols, sum_cols, start=2):
    ws.cell(row=row, column=1, value="TOTAL")
    for ci in sum_cols:
        cl = get_column_letter(ci)
        ws.cell(row=row, column=ci).value = f'=SUM({cl}{start}:{cl}{row - 1})'
        ws.cell(row=row, column=ci).number_format = MONEY_FMT
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.border = TOTAL_FONT, TOTAL_FILL, BORDER


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_entries(wb, data):
    ws = wb.active
    ws.title = "Splitwise Entries"
    payer = data["payer"]
    others = [r for r in data["roommates"] if r != payer]
    owes_hdrs = [f"{r} Owes" for r in others]
    headers = ["Date", "Merchant", "Item", "Total Price", "Split Between",
               "Per Person"] + owes_hdrs + [f"{payer} Share"]
    ws.append(headers)
    ncols = len(headers)

    for e in data["splitwise_entries"]:
        row = [e["date"], e["merchant"], e["item"], e["price"],
               e["split_between"], e["per_person"]]
        for r in others:
            row.append(e.get("owes", {}).get(r, 0))
        row.append(e.get("payer_share", 0))
        ws.append(row)

    last = len(data["splitwise_entries"]) + 1
    tr = last + 1
    money = {4, 6} | set(range(7, ncols + 1))
    total_row(ws, tr, ncols, list(range(7, ncols + 1)) + [4])
    hdr(ws, ncols)
    style_data(ws, 2, last, ncols, money_cols=money)
    auto_width(ws, ncols)
    return tr, others


def build_summary(wb, data, tr, others):
    ws = wb.create_sheet("Per Person Summary")
    ws.append(["Person", f"Total Owed to {data['payer']}", "Breakdown", "Status"])
    col_map = {r: get_column_letter(7 + i) for i, r in enumerate(others)}

    for p in data["person_summary"]:
        ws.append([p["name"], None, p["breakdown"], p.get("status", "Pending")])
        r = ws.max_row
        ws.cell(row=r, column=2).value = f"='Splitwise Entries'!{col_map[p['name']]}{tr}"

    tr2 = ws.max_row + 1
    ws.cell(row=tr2, column=1, value="TOTAL OWED")
    ws.cell(row=tr2, column=2).value = f'=SUM(B2:B{tr2 - 1})'
    ws.cell(row=tr2, column=4, value="-")
    for c in range(1, 5):
        cell = ws.cell(row=tr2, column=c)
        cell.font, cell.fill, cell.border = TOTAL_FONT, TOTAL_FILL, BORDER
    hdr(ws, 4)
    style_data(ws, 2, tr2 - 1, 4, money_cols={2})
    ws.cell(row=tr2, column=2).number_format = MONEY_FMT
    auto_width(ws, 4)
    ws.column_dimensions['C'].width = 50


def build_already(wb, data):
    ws = wb.create_sheet("Already on Splitwise")
    ws.append(["Date", "Merchant", "Amount", "Status"])
    for e in data.get("already_on_splitwise", []):
        ws.append([e["date"], e["merchant"], e["amount"], "Already on Splitwise"])
    tr = ws.max_row + 1
    total_row(ws, tr, 4, [3])
    hdr(ws, 4)
    style_data(ws, 2, tr - 1, 4, money_cols={3})
    auto_width(ws, 4)


def build_personal(wb, data):
    ws = wb.create_sheet("Personal Expenses")
    ws.append(["Date", "Merchant", "Amount", "Category"])
    for e in data.get("personal_expenses", []):
        ws.append([e["date"], e["merchant"], e["amount"], e["category"]])
    tr = ws.max_row + 1
    total_row(ws, tr, 4, [3])
    hdr(ws, 4)
    style_data(ws, 2, tr - 1, 4, money_cols={3})
    auto_width(ws, 4)


def build_tbd(wb, data):
    ws = wb.create_sheet("TBD - Pending")
    ws.append(["Date", "Merchant", "Amount", "Status"])
    for e in data.get("tbd", []):
        ws.append([e["date"], e["merchant"], e["amount"], "Pending confirmation"])
    hdr(ws, 4)
    if ws.max_row > 1:
        style_data(ws, 2, ws.max_row, 4, money_cols={3})
    auto_width(ws, 4)


def build_datewise(wb, data):
    ws = wb.create_sheet("Date-wise for Splitwise")
    cur = 1
    for txn in data.get("datewise_transactions", []):
        # Title bar
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=4)
        c = ws.cell(row=cur, column=1, value=f"{txn['date']}  \u2014  {txn['merchant']}")
        c.font, c.fill = TITLE_FONT, TITLE_FILL
        c.alignment = Alignment(horizontal='center')
        for col in range(1, 5):
            ws.cell(row=cur, column=col).fill = TITLE_FILL
            ws.cell(row=cur, column=col).border = BORDER
        cur += 1

        # Info row
        for col, val in [(1, "Paid by:"), (2, txn.get("paid_by", data["payer"])),
                         (3, "Bill Total:"), (4, txn["bill_total"])]:
            cell = ws.cell(row=cur, column=col, value=val)
            cell.font = SUBTITLE_FONT
            cell.border = BORDER
        ws.cell(row=cur, column=4).number_format = MONEY_FMT
        cur += 1

        # Item headers
        for col, h in enumerate(["Item", "Price", "Shared With", "Per Person"], 1):
            cell = ws.cell(row=cur, column=col, value=h)
            cell.font, cell.fill = ITEM_HDR_FONT, ITEM_HDR_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER
        cur += 1

        # Items
        for item in txn["items"]:
            ws.cell(row=cur, column=1, value=item["name"]).border = BORDER
            for col, key in [(2, "price"), (4, "per_person")]:
                cell = ws.cell(row=cur, column=col, value=item[key])
                cell.border, cell.number_format = BORDER, MONEY_FMT
                cell.alignment = Alignment(horizontal='right')
            ws.cell(row=cur, column=3, value=item["shared_with"]).border = BORDER
            cur += 1

        # Subtotal
        ws.cell(row=cur, column=1, value="TOTAL").font = SUB_FONT
        tc = ws.cell(row=cur, column=2, value=txn["bill_total"])
        tc.font, tc.number_format = SUB_FONT, MONEY_FMT
        for col in range(1, 5):
            ws.cell(row=cur, column=col).fill = SUB_FILL
            ws.cell(row=cur, column=col).border = BORDER
        cur += 2  # blank separator

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 15


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    data = json.load(sys.stdin)
    output = os.path.expanduser(data["output_path"])
    wb = openpyxl.Workbook()
    tr, others = build_entries(wb, data)
    build_summary(wb, data, tr, others)
    build_already(wb, data)
    build_personal(wb, data)
    build_tbd(wb, data)
    build_datewise(wb, data)
    wb.save(output)
    print(json.dumps({"status": "success", "path": output}))


if __name__ == "__main__":
    main()
